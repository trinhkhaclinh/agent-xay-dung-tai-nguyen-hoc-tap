# -*- coding: utf-8 -*-
"""Render Quiz Quizizz: quiz_*.json -> Câu hỏi Quiziz session/<filename>.xlsx

Lược đồ cột CHÍNH XÁC (xem knowledge/format-specs/04-quiz.md).
"""
import os
import sys

import openpyxl
from openpyxl.styles import Font

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from lib.io_utils import load_json, out_path

HEADERS = [
    "question_content",
    "answer_1", "explanation_answer_1",
    "answer_2", "explanation_answer_2",
    "answer_3", "explanation_answer_3",
    "answer_4", "explanation_answer_4",
    "isCorrect", "difficulty", "category",
]


def _validate(q, i):
    ans = q.get("answers", [])
    exp = q.get("explanations", [])
    if len(ans) != 4:
        raise ValueError("Câu %d: cần đúng 4 đáp án, có %d" % (i, len(ans)))
    if len(exp) != 4:
        raise ValueError("Câu %d: cần đúng 4 giải thích, có %d" % (i, len(exp)))
    c = q.get("correct")
    if c not in (1, 2, 3, 4):
        raise ValueError("Câu %d: isCorrect phải ∈ {1,2,3,4}, nhận %r" % (i, c))
    if not str(q.get("category", "")).strip():
        raise ValueError("Câu %d: thiếu category" % i)


def render(spec, session_dir):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = spec.get("sheet_name", "Quiz")[:31]
    ws.append(HEADERS)
    for c in ws[1]:
        c.font = Font(bold=True)
    for i, q in enumerate(spec.get("questions", []), 1):
        _validate(q, i)
        a = q["answers"]
        e = q["explanations"]
        ws.append([
            q.get("q", ""),
            a[0], e[0], a[1], e[1], a[2], e[2], a[3], e[3],
            int(q["correct"]),
            q.get("difficulty", 4),
            q.get("category", ""),
        ])
    # độ rộng cột dễ đọc
    widths = [50, 22, 30, 22, 30, 22, 30, 22, 30, 9, 10, 12]
    for idx, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = w
    path = out_path(session_dir, spec.get("subdir", "Câu hỏi Quiziz session"),
                    spec["filename"], ".xlsx")
    wb.save(path)
    return path


def main():
    if len(sys.argv) < 3:
        print("Cách dùng: python tools/render_quiz.py <spec.json> <session_dir>")
        sys.exit(1)
    spec = load_json(sys.argv[1])
    print("=>", render(spec, sys.argv[2]))


if __name__ == "__main__":
    main()
